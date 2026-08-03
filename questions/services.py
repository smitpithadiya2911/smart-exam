import random
import openpyxl
from django.http import HttpResponse
from .models import Question

class QuestionService:
    @staticmethod
    def generate_random_questions(subject_id, count, difficulty=None, question_type=None):
        """Reusable service to sample random questions from the Question Bank."""
        from subjects.models import Subject
        subject_obj = Subject.objects.filter(id=subject_id).first()
        
        qs = Question.objects.filter(subject_id=subject_id)
        if difficulty:
            qs = qs.filter(difficulty=difficulty)
        if question_type:
            qs = qs.filter(question_type=question_type)
        
        existing_count = qs.count()
        if existing_count < count and subject_obj:
            needed = count - existing_count
            for i in range(1, needed + 1):
                q_num = existing_count + i
                Question.objects.create(
                    subject=subject_obj,
                    question_type=Question.Type.MCQ,
                    prompt_text=f"Sample Assessment Question #{q_num} for {subject_obj.name}: Which of the following statements is correct?",
                    option_a=f"Option A: Primary feature for concept #{q_num}",
                    option_b=f"Option B: Standard implementation rule #{q_num}",
                    option_c=f"Option C: Fundamental property of {subject_obj.code}",
                    option_d=f"Option D: Advanced analytical condition #{q_num}",
                    correct_answer=random.choice(['A', 'B', 'C', 'D']),
                    explanation=f"Explanation for Question #{q_num}: Option choice reflects core subject curriculum guidelines.",
                    marks=5.0,
                    difficulty=Question.Difficulty.MEDIUM
                )
            qs = Question.objects.filter(subject_id=subject_id)

        question_ids = list(qs.values_list('id', flat=True))
        if len(question_ids) <= count:
            selected_ids = question_ids
        else:
            selected_ids = random.sample(question_ids, count)
        
        return Question.objects.filter(id__in=selected_ids)

    @staticmethod
    def export_questions_to_excel(subject_id=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Question Bank"

        headers = [
            'ID', 'Subject Code', 'Type', 'Chapter', 'Topic', 'Marks', 'Difficulty',
            'Prompt', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Answer', 'Explanation', 'Tags'
        ]
        ws.append(headers)

        qs = Question.objects.all().select_related('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        for q in qs:
            ws.append([
                q.id, q.subject.code, q.question_type, q.chapter or '', q.topic or '',
                float(q.marks), q.difficulty, q.prompt_text, q.option_a or '', q.option_b or '',
                q.option_c or '', q.option_d or '', q.correct_answer, q.explanation or '', q.tags or ''
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=question_bank_export.xlsx'
        wb.save(response)
        return response

    @staticmethod
    def import_questions_from_excel(file_obj, subject_obj):
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active
        created_count = 0

        # Skip header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[7]: # prompt_text is required
                continue
            
            q_type = str(row[2]).upper() if row[2] else 'MCQ'
            if q_type not in Question.Type.values:
                q_type = 'MCQ'

            diff = str(row[6]).upper() if row[6] else 'MEDIUM'
            if diff not in Question.Difficulty.values:
                diff = 'MEDIUM'

            Question.objects.create(
                subject=subject_obj,
                question_type=q_type,
                chapter=str(row[3]) if row[3] else '',
                topic=str(row[4]) if row[4] else '',
                marks=float(row[5]) if row[5] else 1.0,
                difficulty=diff,
                prompt_text=str(row[7]),
                option_a=str(row[8]) if row[8] else '',
                option_b=str(row[9]) if row[9] else '',
                option_c=str(row[10]) if row[10] else '',
                option_d=str(row[11]) if row[11] else '',
                correct_answer=str(row[12]) if row[12] else 'A',
                explanation=str(row[13]) if row[13] else '',
                tags=str(row[14]) if row[14] else ''
            )
            created_count += 1

        return created_count

    @staticmethod
    def import_questions_from_pdf(file_obj, subject_obj):
        """Extract and parse questions from a PDF file."""
        import pypdf
        import re
        from .models import Question

        reader = pypdf.PdfReader(file_obj)
        full_text = ""
        for page in reader.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"

        # Try to find a clear heading for answer key section at the end of the text
        header_patterns = [
            r'^\s*(?:[\*\#]+)?\s*(?:correct\s+)?ans(?:wers?)?\s*keys?(?:\s*:)?\s*(?:[\*\#]+)?\s*$',
            r'^\s*(?:[\*\#]+)?\s*(?:correct\s+)?ans(?:wers?)?\s*sheet(?:\s*:)?\s*(?:[\*\#]+)?\s*$',
            r'^\s*(?:[\*\#]+)?\s*(?:correct\s+)?ans(?:wers?)(?:\s*:)?\s*(?:[\*\#]+)?\s*$',
            r'^\s*(?:[\*\#]+)?\s*keys?(?:\s*:)?\s*(?:[\*\#]+)?\s*$'
        ]
        
        question_text = full_text
        answer_key_text = None
        
        for pattern in header_patterns:
            matches = list(re.finditer(pattern, full_text, re.IGNORECASE | re.MULTILINE))
            if matches:
                last_match = matches[-1]
                question_text = full_text[:last_match.start()]
                answer_key_text = full_text[last_match.end():]
                break

        # Split text into lines
        lines = question_text.split('\n')
        parsed_questions = []
        current_q = None

        # Regex patterns
        # 1. Matches question number, e.g. "1. ", "Q1. ", "Question 1:", "1) "
        q_pattern = re.compile(r'^\s*(?:Q|Question)?\s*(\d+)[\s\.\-\):\]]+\s*(.*)', re.IGNORECASE)
        
        # 2. Matches options A, B, C, D (can match multiple on the same line)
        opt_pattern = re.compile(r'((?:\b[A-D][\.\-\):]|\([A-D]\)|\[[A-D]\]))\s*(.*?)(?=\s*(?:\b[A-D][\.\-\):]|\([A-D]\)|\[[A-D]\])|$)', re.IGNORECASE)
        
        # 3. Matches answers, e.g. "Answer: A", "Correct Answer: B", "Ans: C", "Correct: D"
        ans_pattern = re.compile(r'^\s*(?:correct\s+)?ans(?:wer)?[\s\.\-\:]+\s*(.*)', re.IGNORECASE)
        
        # 4. Matches explanation, e.g. "Explanation: ...", "Explain: ..."
        exp_pattern = re.compile(r'^\s*explanation[\s\.\-\:]+\s*(.*)', re.IGNORECASE)

        for line in lines:
            line_str = line.strip()
            if not line_str:
                continue

            # Clean markdown formatting often generated by AI (e.g. ChatGPT, Gemini)
            line_str = re.sub(r'\*\*(.*?)\*\*', r'\1', line_str)
            line_str = re.sub(r'\*(.*?)\*', r'\1', line_str)
            line_str = re.sub(r'^#+\s*', '', line_str)

            # Check if this line starts a new question
            q_match = q_pattern.match(line_str)
            if q_match:
                if current_q and current_q.get('prompt_text'):
                    parsed_questions.append(current_q)
                
                prompt = q_match.group(2).strip()
                current_q = {
                    'num': int(q_match.group(1)),
                    'prompt_text': prompt,
                    'option_a': '',
                    'option_b': '',
                    'option_c': '',
                    'option_d': '',
                    'correct_answer': None,
                    'explanation': '',
                }

                # Check if options are on the same line as the question
                opt_matches = opt_pattern.findall(prompt)
                if opt_matches:
                    for opt_marker, option_text in opt_matches:
                        letter_match = re.search(r'([A-D])', opt_marker, re.IGNORECASE)
                        if letter_match:
                            letter = letter_match.group(1).upper()
                            current_q[f'option_{letter.lower()}'] = option_text.strip()
                    
                    # Remove options from prompt_text
                    first_opt = re.search(r'((?:\b[A-D][\.\-\):]|\([A-D]\)|\[[A-D]\]))\s*', prompt, re.IGNORECASE)
                    if first_opt:
                        current_q['prompt_text'] = prompt[:first_opt.start()].strip()
                continue

            # Check if this line contains options (A, B, C, or D)
            opt_matches = opt_pattern.findall(line_str)
            if opt_matches and current_q:
                for opt_marker, option_text in opt_matches:
                    letter_match = re.search(r'([A-D])', opt_marker, re.IGNORECASE)
                    if letter_match:
                        letter = letter_match.group(1).upper()
                        current_q[f'option_{letter.lower()}'] = option_text.strip()
                continue


            # Check if this line is an answer
            ans_match = ans_pattern.match(line_str)
            if ans_match and current_q:
                ans_text = ans_match.group(1).strip()
                letter_match = re.search(r'\b([A-D])\b', ans_text, re.IGNORECASE)
                if letter_match:
                    current_q['correct_answer'] = letter_match.group(1).upper()
                else:
                    current_q['correct_answer'] = ans_text[:50].strip().upper()
                continue

            # Check if this line is an explanation
            exp_match = exp_pattern.match(line_str)
            if exp_match and current_q:
                current_q['explanation'] = exp_match.group(1).strip()
                continue

            # If it doesn't match anything, it's a continuation of previous section
            if current_q:
                if not current_q['option_a']:
                    current_q['prompt_text'] += " " + line_str
                elif current_q['explanation']:
                    current_q['explanation'] += " " + line_str
                elif current_q['option_d']:
                    current_q['option_d'] += " " + line_str
                elif current_q['option_c']:
                    current_q['option_c'] += " " + line_str
                elif current_q['option_b']:
                    current_q['option_b'] += " " + line_str
                elif current_q['option_a']:
                    current_q['option_a'] += " " + line_str

        # Add the final question
        if current_q and current_q.get('prompt_text'):
            parsed_questions.append(current_q)

        # Parse answers from the answer key section if we have one or if fallback is needed
        answer_key = {}
        if not answer_key_text and parsed_questions:
            last_q = parsed_questions[-1]
            last_text = last_q.get('option_d') or last_q.get('option_c') or last_q.get('option_b') or last_q.get('option_a') or last_q.get('prompt_text')
            if last_text:
                pos = full_text.rfind(last_text)
                if pos != -1:
                    answer_key_text = full_text[pos + len(last_text):]

        if answer_key_text:
            ans_pat = re.compile(
                r'\b(\d+)\s*[\.\-\)\]\s:\(\[]+\s*(?:ans(?:wer)?|correct)?\s*[\.\-\)\]\s:\(\[]*\s*([A-D])\b', 
                re.IGNORECASE
            )
            for q_num_str, ans_letter in ans_pat.findall(answer_key_text):
                answer_key[int(q_num_str)] = ans_letter.upper()

        # Save to database and return
        saved_questions = []
        for q in parsed_questions:
            if not q['option_a'] or not q['option_b']:
                continue

            correct_ans = q.get('correct_answer')
            if not correct_ans:
                q_num = q.get('num')
                if q_num in answer_key:
                    correct_ans = answer_key[q_num]
                else:
                    correct_ans = 'A' # fallback
            
            db_q = Question.objects.create(
                subject=subject_obj,
                question_type=Question.Type.MCQ,
                prompt_text=q['prompt_text'],
                option_a=q['option_a'],
                option_b=q['option_b'],
                option_c=q['option_c'],
                option_d=q['option_d'],
                correct_answer=correct_ans,
                explanation=q['explanation'],
                marks=5.0, # match default sampler weight
                difficulty=Question.Difficulty.MEDIUM
            )
            saved_questions.append(db_q)

        from questions.models import Question as QuestionModel
        return QuestionModel.objects.filter(id__in=[sq.id for sq in saved_questions])
